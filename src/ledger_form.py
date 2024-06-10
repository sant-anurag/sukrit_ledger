import tkinter as tk
from app_defines import *
from app_defines import *
from app_common import *
from init_database import *
from app_thread import *
import os
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
import tkinter as tk
from tkinter import W,E,messagebox
from PIL import ImageGrab
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
from pathlib import Path


class LedgerForm(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Ledger Form")
        self.geometry("1000x1000")  # Set the window size to 600x600 pixels
        self.configure(bg='lightblue')  # Set background color to light blue
        self.previous_balance_entry = tk.Entry(state='readonly')
        self.total_balance_entry = tk.Entry(state='readonly')
        # Create a container frame to hold all the widgets
        container = tk.Frame(self, bg='lightblue')
        container.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        self.heading_label = tk.Label(container, text="Ledger Form", bg='lightblue', font=("Helvetica", 16, "bold"))
        self.heading_label.grid(row=0, columnspan=2, pady=10)
        
        # Create labels and entry fields for each ledger field
        self.date_label = tk.Label(container, text="Date:", bg='lightblue')
        self.date_entry = tk.Entry(container)
        
        self.party_name_label = tk.Label(container, text="Party's Name:", bg='lightblue')
        self.party_name_entry = tk.Entry(container)
        
        self.location_label = tk.Label(container, text="Location:", bg='lightblue')
        self.location_entry = tk.Entry(container)

        self.party_name_entry.bind("<FocusOut>", self.load_previous_balance)
        
        self.invoice_no_label = tk.Label(container, text="Invoice No:", bg='lightblue')
        self.invoice_no_entry = tk.Entry(container)
        
        self.billing_amount_label = tk.Label(container, text="Billing Amount:", bg='lightblue')
        self.billing_amount_entry = tk.Entry(container)
        
        self.amount_received_label = tk.Label(container, text="Amount Received:", bg='lightblue')
        self.amount_received_entry = tk.Entry(container)

        self.receipt_ref_no_label = tk.Label(container, text="MR No.:", bg='lightblue')
        self.receipt_ref_no_entry = tk.Entry(container)
        
        self.amount_balance_label = tk.Label(container, text="Amount Balance:", bg='lightblue')
        self.amount_balance_entry = tk.Entry(container, state='readonly')
        
        self.previous_balance_label = tk.Label(container, text="Previous Balance:", bg='lightblue')
        self.previous_balance_entry = tk.Entry(container, state='readonly')

        self.total_balance_label = tk.Label(container, text="Total Balance:", bg='lightblue')
        self.total_balance_entry = tk.Entry(container, state='readonly')
        
        self.remarks_label = tk.Label(container, text="Remarks:", bg='lightblue')
        self.remarks_entry = tk.Entry(container)
        
        # Create a submit button to save the ledger entry
        self.submit_button = tk.Button(container, text="Submit", command=self.submit_ledger)
        self.statement_button = tk.Button(container, text="Generate Statement", command=self.generate_statement)

        self.billing_amount_entry.bind("<KeyRelease>", self.update_amount_balance)
        self.amount_received_entry.bind("<KeyRelease>", self.update_amount_balance)

        

        # Arrange the widgets using grid geometry manager
        self.date_label.grid(row=1, column=0, sticky=tk.E, pady=5, padx=10)
        self.date_entry.grid(row=1, column=1, pady=5, padx=10)
        self.party_name_label.grid(row=2, column=0, sticky=tk.E, pady=5, padx=10)
        self.party_name_entry.grid(row=2, column=1, pady=5, padx=10)
        self.location_label.grid(row=3, column=0, sticky=tk.E, pady=5, padx=10)
        self.location_entry.grid(row=3, column=1, pady=5, padx=10)
        self.invoice_no_label.grid(row=4, column=0, sticky=tk.E, pady=5, padx=10)
        self.invoice_no_entry.grid(row=4, column=1, pady=5, padx=10)
        self.billing_amount_label.grid(row=5, column=0, sticky=tk.E, pady=5, padx=10)
        self.billing_amount_entry.grid(row=5, column=1, pady=5, padx=10)
        self.amount_received_label.grid(row=6, column=0, sticky=tk.E, pady=5, padx=10)
        self.amount_received_entry.grid(row=6, column=1, pady=5, padx=10)
        self.receipt_ref_no_label.grid(row=7, column=0, sticky=tk.E, pady=5, padx=10)
        self.receipt_ref_no_entry.grid(row=7, column=1, pady=5, padx=10)
        self.amount_balance_label.grid(row=8, column=0, sticky=tk.E, pady=5, padx=10)
        self.amount_balance_entry.grid(row=8, column=1, pady=5, padx=10)
        self.previous_balance_label.grid(row=9, column=0, sticky=tk.E, pady=5, padx=10)
        self.previous_balance_entry.grid(row=9, column=1, pady=5, padx=10)
        self.total_balance_label.grid(row=10, column=0, sticky=tk.E, pady=5, padx=10)
        self.total_balance_entry.grid(row=10, column=1, pady=5, padx=10)
        self.remarks_label.grid(row=11, column=0, sticky=tk.E, pady=5, padx=10)
        self.remarks_entry.grid(row=11, column=1, pady=5, padx=10)
        
        self.submit_button.grid(row=12, columnspan=2, pady=10)
        self.statement_button.grid(row=13, columnspan=2, pady=10)

        self.load_previous_balance()

    def load_previous_balance(self, event=None):
        party_name_input = self.party_name_entry.get().strip()
        

        if not party_name_input:
            return

        try:
            today = datetime.now()
            year = today.strftime("%Y")
            dirname = "Expanse_Data\\" +  "\\Invoices"
            if not os.path.exists(dirname):
                print("Current year directory is not available , hence building one")
                # os.makedirs(dirname)
            file_path = dirname + "\\ledger_Details.xlsx"
            if os.path.exists(file_path):
                print("exist")
                try:
                   
                    workbook = load_workbook(file_path)
                    
                    sheet = workbook.active
                    
                    # Find the last row with data
                    last_row = sheet.max_row
                    

                    # Initialize latest total balance
                    latest_total_balance = None

                    party_name_found = False
                    for row in range(2, last_row + 1):  # Assuming row 1 has headers
                        party_name = sheet.cell(row=row, column=2).value  # Column 2 for party names
                        if party_name:
                            party_name = party_name.strip()
                            if party_name.lower() == party_name_input.strip().lower():  # Case-insensitive comparison
                                
                                total_pending = sheet.cell(row=row, column=10).value  # Assuming total amount is in column 10
                                latest_total_balance = total_pending
                                
                                party_name_found = True
                                break

                    if not party_name_found:
                        print(f"Party name '{party_name_input}' not found. Adding new entry...")
                        new_row = last_row + 1
                        sheet.cell(row=new_row, column=2, value=party_name_input)  # Add party name to column 2
                        # Add any additional information for the new entry
                        # For example, if the total pending amount needs to be added:
                        # sheet.cell(row=new_row, column=10, value=<total_pending_amount>)
                        # Set latest_total_balance to some default value for new entry
                        latest_total_balance = 0.0


                    # Update the GUI entries with the captured balance
                    self.previous_balance_entry.config(state='normal')
                    self.previous_balance_entry.delete(0, tk.END)
                    self.previous_balance_entry.insert(0, str(latest_total_balance))
                    self.previous_balance_entry.config(state='readonly')

                    self.total_balance_entry.config(state='normal')
                    self.total_balance_entry.delete(0, tk.END)
                    self.total_balance_entry.insert(0, str(latest_total_balance))
                    self.total_balance_entry.config(state='readonly')

                except Exception as e:
                    print("Error accessing file:", e)
                    messagebox.showerror("Error", f"Error accessing file: {e}")

            else:
                self.previous_balance_entry.config(state='normal')
                self.previous_balance_entry.delete(0, tk.END)
                self.previous_balance_entry.insert(0, "0.0")
                self.previous_balance_entry.config(state='readonly')

                self.total_balance_entry.config(state='normal')
                self.total_balance_entry.delete(0, tk.END)
                self.total_balance_entry.insert(0, "0.0")
                self.total_balance_entry.config(state='readonly')

        except Exception as e:
            messagebox.showerror("Error", f"Error occurred while loading previous balance: {e}")


    def update_amount_balance(self, event):
        billing_amount = self.billing_amount_entry.get()
        amount_received = self.amount_received_entry.get()
        
        try:
            billing_amount_float = float(billing_amount)
            amount_received_float = float(amount_received)
            amount_balance = billing_amount_float - amount_received_float

            previous_balance = float(self.previous_balance_entry.get())
            total_balance_pending = previous_balance + amount_balance

            self.amount_balance_entry.config(state='normal')
            self.amount_balance_entry.delete(0, tk.END)
            self.amount_balance_entry.insert(0, str(amount_balance))
            self.amount_balance_entry.config(state='readonly')

            self.total_balance_entry.config(state='normal')
            self.total_balance_entry.delete(0, tk.END)
            self.total_balance_entry.insert(0, str(total_balance_pending))
            self.total_balance_entry.config(state='readonly')

        except ValueError:
            self.amount_balance_entry.config(state='normal')
            self.amount_balance_entry.delete(0, tk.END)
            self.amount_balance_entry.insert(0, "Invalid Input")
            self.amount_balance_entry.config(state='readonly')

    def clear_form(self):
    # Clear all entry fields
        self.date_entry.delete(0, tk.END)
        self.party_name_entry.delete(0, tk.END)
        self.location_entry.delete(0, tk.END)
        self.invoice_no_entry.delete(0, tk.END)
        self.billing_amount_entry.delete(0, tk.END)
        self.amount_received_entry.delete(0, tk.END)
        self.receipt_ref_no_entry.delete(0, tk.END)
        self.amount_balance_entry.delete(0, tk.END)
        self.previous_balance_entry.delete(0, tk.END)
        self.total_balance_entry.delete(0, tk.END)
        self.remarks_entry.delete(0, tk.END)

    def submit_ledger(self):
        date = self.date_entry.get()
        party_name = self.party_name_entry.get()
        location = self.location_entry.get()
        invoice_no = self.invoice_no_entry.get()
        billing_amount = self.billing_amount_entry.get()
        mr_no = self.receipt_ref_no_entry.get()
        amount_received = self.amount_received_entry.get()
        amount_balance = self.amount_balance_entry.get()
        previous_balance = self.previous_balance_entry.get()
        total_balance_pending = self.total_balance_entry.get()
        remarks = self.remarks_entry.get()

        try:
            billing_amount_float = float(billing_amount)
            amount_received_float = float(amount_received)
            amount_balance_float = billing_amount_float - amount_received_float
            
            previous_balance_float = float(previous_balance)
            total_balance_pending_float = previous_balance_float + amount_balance_float
            today = datetime.now()
            year = today.strftime("%Y")
            # Construct the file path for the Excel file
            dirname = "Expanse_Data\\"  + "\\Invoices"
            if not os.path.exists(dirname):
                print("Current year directory is not available , hence building one")
                os.makedirs(dirname)
            file_path = dirname + "\\ledger_Details.xlsx"
            if not os.path.isfile(file_path):
                wb = openpyxl.Workbook()
                sheet = wb.active
            
            # Load or create the workbook
            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Date", "Party Name", "Location", "Invoice No", "Billing Amount", "Amount Received", "MR No","Amount Balance","Previous Balance" ,"Total Balance", "Remarks"])

            sheet = workbook.active
            sheet.append([date, party_name, location, invoice_no, billing_amount_float, amount_received_float , mr_no, amount_balance_float,previous_balance, total_balance_pending_float, remarks])

            workbook.save(file_path)

            self.previous_balance_entry.config(state='normal')
            self.previous_balance_entry.delete(0, tk.END)
            self.previous_balance_entry.insert(0, str(total_balance_pending_float))
            self.previous_balance_entry.config(state='readonly')

            self.amount_balance_entry.config(state='normal')
            self.amount_balance_entry.delete(0, tk.END)
            self.amount_balance_entry.config(state='readonly')

            self.total_balance_entry.config(state='normal')
            self.total_balance_entry.delete(0, tk.END)
            self.total_balance_entry.config(state='readonly')

            self.remarks_entry.delete(0, tk.END)

            messagebox.showinfo("Success", "Ledger entry submitted successfully!")
            self.clear_form()

        except ValueError:
            messagebox.showerror("Input Error", "Billing Amount and Amount Received should be numbers.")
        except Exception as e:
            messagebox.showerror("Error", f"Error occurred: {e}")
            
    def generate_statement(self):
        party_name = self.party_name_entry.get()
        
        if not party_name:
            messagebox.showerror("Input Error", "Please enter the Party's Name.")
            return

        # Construct the file path
        today = datetime.now()
        year = today.strftime("%Y")
        directory = os.path.join("../Expanse_Data", year, "Invoices")
        file_path = os.path.join(directory, "ledger_Details.xlsx")

        if not os.path.exists(file_path):
            messagebox.showerror("File Error", "No ledger data found.")
            return

        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            records = []

            for row in sheet.iter_rows(values_only=True):
                if row[1] == party_name:
                    records.append(row)
            
            if records:
                self.show_records(records)
            else:
                messagebox.showinfo("No Records", "No records found for the given party.")
        
        except InvalidFileException:
            messagebox.showerror("File Error", f"The file '{file_path}' is not a valid Excel file. Please check the file format.")
        except Exception as e:
            messagebox.showerror("Error", f"Error occurred: {e}")

    def show_records(self, records):
        # Get the current month and year
        today = datetime.now()
        year = today.strftime("%Y")
        month = today.strftime("%B")

        # Create the statement
        statement = f"Statement of Invoice for {month}, {year}"

        # Create a window to display the records
        self.records_window = Toplevel(self)
        self.records_window.title(f"Statement of Invoice for {month}, {year}")
        self.records_window.geometry("600x400")
        self.records_window.configure(bg='lightblue')

        # Define column names
        columns = [
            "Date",
            "Party's Name",
            "Location",
            "Invoice No",
            "Billing Amount",
            "Amount Received",
            "MR No"
            "Amount Balance",
            "Previous Balance",
            "Total Balance",
            "Remarks"
        ]

        # Create labels for column names
        for i, column_name in enumerate(columns):
            Label(self.records_window, text=column_name, font=("Helvetica", 10, "bold"), bg='lightblue').grid(row=0, column=i, padx=10, pady=5)

        # Display the records
        for i, record in enumerate(records):
            for j, value in enumerate(record):
                Label(self.records_window, text=value, bg='lightblue').grid(row=i+1, column=j, padx=10, pady=5)

        # Create a print button
        print_button = Button( self.records_window,text="Print", command=lambda: self.print_window(self.records_window))
        print_button.grid(row=len(records) + 2, columnspan=len(columns), pady=10)

    def print_window(self):
        if self.records_window:
            try:
                # Capture the content of the records_window as an image
                x = self.records_window.winfo_rootx()
                y = self.records_window.winfo_rooty()
                width = self.records_window.winfo_width()
                height = self.records_window.winfo_height()
                x1 = x + width
                y1 = y + height
                ImageGrab.grab().crop((x, y, x1, y1)).print()
            except Exception as e:
                messagebox.showerror("Printing Error", f"An error occurred while printing: {e}")
        else:
            messagebox.showerror("Printing Error", "No content to print.")

    def myfunction(self, mycanvas, event):
        mycanvas.configure(scrollregion=mycanvas.bbox("all"), width=725, height=407)

    def closepage(self, src_file, starting_index, gaushala_account_statement_window):
        print("closepage :", src_file)
        gaushala_account_statement_window.destroy()
        # erase the written records in template sheet
        # executed only if
        wb_template = openpyxl.load_workbook(src_file)
        template_sheet = wb_template.active

        for rows in range(15, starting_index + 1):
            for columns in range(1, 7):
                template_sheet.cell(row=rows, column=columns).value = ""

        wb_template.save(src_file)

    def prepare_account_statement_Excel(self, gaushala_account_statement_window, printBtn, cal_dateFrom,
                                        cal_toDate, viewbyMonth_monthTxt, viewbymonth_yearTxt,
                                        viewbyYear_yearTxt, var, viewPDF, infoLabel, cancelbtn):
        print("prepare_main_account_statement_Excel --> var:", var.get())

        viewPDF.configure(state=DISABLED, bg="light grey")
        printBtn.configure(state=DISABLED, bg="light grey")
        if var.get() == 1:
            dateTimeObj_From = cal_dateFrom.get_date()
            from_Date = dateTimeObj_From.strftime("%Y-%m-%d")
            dateTimeObj_To = cal_toDate.get_date()
            to_Date = dateTimeObj_To.strftime("%Y-%m-%d")
            fromDate = self.dateTimeOp.prepare_dateFromString(from_Date)
            toDate = self.dateTimeOp.prepare_dateFromString(to_Date)
        elif var.get() == 2:
            noOfDays, month_number = self.dateTimeOp.calculateNoOfDaysInMonth(viewbyMonth_monthTxt.get(),
                                                                              viewbymonth_yearTxt.get())
            fromDate, toDate = self.dateTimeOp.getFromAndToDates_Account_Statement(month_number,
                                                                                   viewbymonth_yearTxt.get(),
                                                                                   noOfDays)
        else:
            print("Requested year is :", viewbyYear_yearTxt.get())
            noOfDays = self.dateTimeOp.calculateNoOfDaysInYear(viewbyYear_yearTxt.get())
            fromDate, toDate = self.dateTimeOp.getFromAndToDates_Account_Statement(1, viewbyYear_yearTxt.get(),
                                                                                   noOfDays)

        from_year = fromDate.strftime("%Y")
        to_year = toDate.strftime("%Y")
        to_month = toDate.strftime('%m')
        today_date = datetime.now()
        formatted_date = today_date.strftime("%Y-%m-%d")
        currentDate = self.dateTimeOp.prepare_dateFromString(formatted_date)

        current_year = currentDate.strftime("%Y")
        print("From Year :", from_year, " To Year :", to_year)
        bDateConditionsValid = True
        if fromDate > toDate:
            error_info = "From date cannot be grater than to date !!!"
            bDateConditionsValid = False
        elif ((toDate > currentDate) or (fromDate > currentDate)) and \
                ((var.get() == 1) or (var.get() == 2)):

            if var.get() == 2:
                current_month = datetime.today().month
                current_year = datetime.today().year
                print("to_year : ", to_year, "current_year :", current_year, "to_month :", to_month, "current_month :",
                      current_month)
                if int(to_year) > int(current_year) or int(to_month) > int(current_month):
                    error_info = "Year/Month can not be future !!!"
                    bDateConditionsValid = False
                if int(to_year) == int(current_year) and int(to_month) == int(current_month):
                    bDateConditionsValid = True
            else:
                error_info = "From/To Date cannot be greater than today!!!"
                bDateConditionsValid = False
        elif ((toDate - fromDate).days > 180) and (var.get() == 1):
            error_info = "Statements can be generated for maximum of 180 days !!! "
            bDateConditionsValid = False
        else:
            # check if the selected years are less than or equal to current date ,
            # but no database exists for them
            dir_name = "..\\Expanse_Data\\" + str(from_year)
            if not os.path.exists(dir_name):
                error_info = "Database does not exists for " + str(from_year) + " Please correct !!!"
                bDateConditionsValid = False
            pass

        # algorithm generates the statement when from and start date has the same year
        # same current year directory needs to be referred for these statement generations
        print("bDateConditionsValid :", bDateConditionsValid)
        if bDateConditionsValid:
            if from_year == self.obj_commonUtil.getCurrentYearFolderName() and \
                    to_year == self.obj_commonUtil.getCurrentYearFolderName():
                bAlike_seva = True
                print("This is current year transaction")

                path_seva_sheet = InitDatabase.getInstance().get_gaushala_transaction_database_name()  # Main Transaction Database
                InitDatabase.getInstance().initilize_sorted_Transaction_database()
                sorted_seva_sheet = InitDatabase.getInstance().get_sorted_transaction_database_name()

                if bAlike_seva:
                    print("path_seva_sheet :", path_seva_sheet)
                    wb_obj = openpyxl.load_workbook(path_seva_sheet)
                    wb_sorted_obj = openpyxl.load_workbook(sorted_seva_sheet)
                    sheet_obj = wb_obj.active
                    sheet_sorted_obj = wb_sorted_obj.active
                    total_records = self.obj_commonUtil.totalrecords_excelDataBase(path_seva_sheet)
                    if total_records > 0:
                        sort_sheet_index = 2
                        print("Total records  in transaction sheet:", total_records)
                        for row_index in range(0, total_records):
                            # critical stock ->stock with quantity is 0 or 1
                            # print("Date from sheet is :", sheet_obj.cell(row=row_index + 2, column=6).value)
                            dateFromTransactionSheet = self.dateTimeOp.prepare_dateFromString(
                                sheet_obj.cell(row=row_index + 2, column=2).value)  # Date of transaction in the sheet
                            print("dateFromMon_DepositSheet :", dateFromTransactionSheet, "fromDate :", fromDate,
                                  " toDate:", toDate)

                            if ((dateFromTransactionSheet > fromDate or dateFromTransactionSheet == fromDate)
                                    and (dateFromTransactionSheet < toDate or dateFromTransactionSheet == toDate)):
                                print("Condition is success")
                                for column_index in range(1, 11):
                                    text_value = str(sheet_obj.cell(row=row_index + 2, column=column_index).value)

                                    sheet_sorted_obj.cell(row=sort_sheet_index, column=column_index).font = Font(size=8,
                                                                                                                 name='Arial',
                                                                                                                 bold=False)
                                    sheet_sorted_obj.cell(row=sort_sheet_index,
                                                          column=column_index).alignment = Alignment(
                                        horizontal='left', vertical='center', wrapText=True)
                                    sheet_sorted_obj.cell(row=sort_sheet_index, column=column_index).value = text_value

                                sort_sheet_index = sort_sheet_index + 1

                        today = date.today()
                        dt_today = today.strftime("%d-%b-%Y")
                        wb_sorted_obj.save(sorted_seva_sheet)
                        print("Sorted transaction sheet created for sorting")
                        self.obj_commonUtil.sortExcelSheetByDate(sorted_seva_sheet, sorted_seva_sheet)

                        now = datetime.now()
                        dt_string = now.strftime("%d_%b_%Y_%H%M%S")
                        currentyear = now.strftime("%Y")
                        destination_file = "..\\Expanse_Data\\" + currentyear + "\\Transaction\\Account_Statement\\Statements\\Gaushala_Account_Statement" + dt_string + ".pdf"
                        # write the  sorted record in statement template
                        template_sheet = "..\\Expanse_Data\\" + currentyear + "\\Transaction\\Account_Statement\\Template\\Transaction_template.xlsx"
                        wb_critical_stock = openpyxl.load_workbook(template_sheet)
                        critical_stock_sheet = wb_critical_stock.active
                        total_sorted_records = self.obj_commonUtil.totalrecords_excelDataBase(sorted_seva_sheet)
                        wb_sort = openpyxl.load_workbook(sorted_seva_sheet)
                        sort_sheet = wb_sort.active
                        dict_index = 1
                        starting_index = 15
                        print("Total sorted records :", total_sorted_records)
                        if total_sorted_records > 0:
                            text_info = "Statement is being generated for Gaushala Accounts.Please wait ...."
                            infoLabel.configure(text=text_info, fg='purple')
                            for row_index in range(1, total_sorted_records + 1):
                                # For the first entry in the account statement , respective credit/debit balance is the
                                # main balance
                                if dict_index == 1:
                                    # if credit column is numeric meaning transaction is credit candidate
                                    if str(sort_sheet.cell(row=row_index + 1, column=3).value).isnumeric():
                                        balance = int(
                                            sort_sheet.cell(row=row_index + 1, column=3).value)  # credit column
                                    else:
                                        # else transaction is debit candidate
                                        balance = int(
                                            sort_sheet.cell(row=row_index + 1, column=4).value)  # debit column
                                else:
                                    if str(sort_sheet.cell(row=row_index + 1,
                                                           column=3).value).isnumeric():  # credit amount is added
                                        balance = balance + int(sort_sheet.cell(row=row_index + 1, column=3).value)
                                    else:  # debit amount is substracted
                                        balance = balance - int(sort_sheet.cell(row=row_index + 1, column=4).value)

                                for column_index in range(1, 7):
                                    if column_index == 1:  # Date
                                        text_value = sort_sheet.cell(row=row_index + 1, column=2).value
                                        text_value = text_value.strftime("%d-%b-%Y")
                                    elif column_index == 2:  # Invoice
                                        text_value = sort_sheet.cell(row=row_index + 1, column=10).value
                                    elif column_index == 3:  # Description
                                        text_value = str(sort_sheet.cell(row=row_index + 1, column=5).value) + "-By " + \
                                                     str(sort_sheet.cell(row=row_index + 1,
                                                                         column=6).value) + "-From " + \
                                                     str(sort_sheet.cell(row=row_index + 1, column=8).value)
                                    elif column_index == 4:  # credit
                                        text_value = sort_sheet.cell(row=row_index + 1, column=3).value
                                    elif column_index == 5:  # debit
                                        text_value = sort_sheet.cell(row=row_index + 1, column=4).value
                                    elif column_index == 6:  # balance
                                        text_value = str(balance)
                                    else:
                                        pass
                                    critical_stock_sheet.cell(row=starting_index, column=column_index).font = Font(
                                        size=8,
                                        name='Arial',
                                        bold=False)
                                    if column_index == 4 or column_index == 5 or column_index == 6:
                                        critical_stock_sheet.cell(row=starting_index,
                                                                  column=column_index).alignment = Alignment(
                                            horizontal='center', vertical='center', wrapText=True)
                                    else:
                                        critical_stock_sheet.cell(row=starting_index,
                                                                  column=column_index).alignment = Alignment(
                                            horizontal='left', vertical='center', wrapText=True)

                                    critical_stock_sheet.cell(row=starting_index,
                                                              column=column_index).value = text_value
                                dict_index = dict_index + 1
                                starting_index = starting_index + 1

                            frdateforstatement = fromDate.strftime("%d-%b-%Y")
                            todateforstatement = toDate.strftime("%d-%b-%Y")
                            critical_stock_sheet.cell(row=3, column=6).value = dt_today
                            critical_stock_sheet.cell(row=5, column=6).value = "Accounts"
                            critical_stock_sheet.cell(row=8, column=6).value = str(balance)
                            critical_stock_sheet.cell(row=9, column=6).value = str(starting_index - 15)
                            critical_stock_sheet.cell(row=10, column=6).value = frdateforstatement
                            critical_stock_sheet.cell(row=11, column=6).value = todateforstatement
                            wb_critical_stock.save(template_sheet)
                            print("File has been saved for template")
                            destination_copy_folder = InitDatabase.getInstance().get_desktop_statement_directory_path()
                            obj_threadClass = myThread(10, "statementThread", 1, template_sheet,
                                                       destination_file, starting_index, viewPDF, printBtn, infoLabel,
                                                       destination_copy_folder)
                            obj_threadClass.start()
                            print_result = partial(self.obj_commonUtil.print_statement_file,
                                                   template_sheet,
                                                   destination_file, starting_index)
                            printBtn.configure(command=print_result)
                            view_result = partial(self.obj_commonUtil.open_statement_file, template_sheet,
                                                  destination_file, starting_index)
                            viewPDF.configure(command=view_result, )

                            cancel_result = partial(self.closepage, template_sheet, starting_index,
                                                    gaushala_account_statement_window)
                            cancelbtn.configure(command=cancel_result)
                        else:
                            text_error = "No records present for Transaction in specified period!!!"
                            infoLabel.configure(text=text_error, fg='red')
                    else:
                        text_error = "No records present for  Transaction"
                        infoLabel.configure(text=text_error, fg='red')
            else:
                # algorithm generates the statement when from and start date has different year or
                # transaction is requested for year other than current year
                # same current year directory needs to be referred for these statement generations
                print(" From year and to year are different other than current year")

                # Since the maximum viewed transaction are only 6 months
                # only 2 year numbers can be considered at max
                # hence same loop with different from and to dates in executed twice
                # this is possible only in case of view by date
                yearDiff = int(to_year) - int(from_year)
                loop_range = yearDiff + 2

                print("Year diff :", yearDiff, " loop_range :", loop_range)
                for year_loop in range(1, loop_range):
                    if var.get() == 1:
                        if year_loop == 1:
                            fDate = fromDate
                            yearOfFromdate = fDate.strftime("%Y")
                            yearFolderToSearch = yearOfFromdate
                            tDate = self.obj_commonUtil.prepare_dateFromString("31" + "-" + "12" + "-" + yearOfFromdate)
                        elif year_loop == 2:
                            yearOfTodate = toDate.strftime("%Y")
                            fDate = self.obj_commonUtil.prepare_dateFromString("1" + "-" + "1" + "-" + yearOfTodate)
                            yearOfTodate = toDate.strftime("%Y")
                            yearFolderToSearch = yearOfTodate
                            tDate = toDate
                        else:
                            pass
                    elif var.get() == 2 or var.get() == 3:
                        fDate = fromDate
                        tDate = toDate
                        yearOfFromdate = fDate.strftime("%Y")
                        yearFolderToSearch = yearOfFromdate
                    else:
                        pass

                    print("fDate :", fDate, "tDate:", tDate, "yearFolderToSearch :", yearFolderToSearch)

                    path_seva_sheet = "..\\Expanse_Data\\" + yearFolderToSearch + "\\Transaction\\Gaushala_Transaction.xlsx"
                    if year_loop == 1:
                        InitDatabase.getInstance().initilize_sorted_Transaction_database()
                        sorted_seva_sheet = InitDatabase.getInstance().get_sorted_transaction_database_name()

                    print("path_seva_sheet :", path_seva_sheet)
                    wb_obj = openpyxl.load_workbook(path_seva_sheet)
                    wb_sorted_obj = openpyxl.load_workbook(sorted_seva_sheet)
                    sheet_obj = wb_obj.active
                    sheet_sorted_obj = wb_sorted_obj.active
                    total_records = self.obj_commonUtil.totalrecords_excelDataBase(path_seva_sheet)
                    if total_records > 0:
                        sort_sheet_index = self.obj_commonUtil.totalrecords_excelDataBase(sorted_seva_sheet) + 2
                        print("Sorted sheet will now start from row:", sort_sheet_index)
                        print("Total records  in transaction sheet:", total_records)
                        for row_index in range(0, total_records):
                            # critical stock ->stock with quantity is 0 or 1
                            # print("Date from sheet is :", sheet_obj.cell(row=row_index + 2, column=6).value)
                            dateFromTransactionSheet = self.dateTimeOp.prepare_dateFromString(
                                sheet_obj.cell(row=row_index + 2, column=2).value)
                            # print("dateFromMon_DepositSheet :", dateFromTransactionSheet, "fromDate :", fromDate, " toDate:", toDate)

                            if ((dateFromTransactionSheet > fromDate or dateFromTransactionSheet == fromDate)
                                    and (dateFromTransactionSheet < toDate or dateFromTransactionSheet == toDate)):
                                # print("Condition is success")
                                for column_index in range(1, 11):
                                    text_value = str(sheet_obj.cell(row=row_index + 2, column=column_index).value)

                                    sheet_sorted_obj.cell(row=sort_sheet_index, column=column_index).font = Font(size=8,
                                                                                                                 name='Arial',
                                                                                                                 bold=False)
                                    sheet_sorted_obj.cell(row=sort_sheet_index,
                                                          column=column_index).alignment = Alignment(
                                        horizontal='left', vertical='center', wrapText=True)
                                    sheet_sorted_obj.cell(row=sort_sheet_index, column=column_index).value = text_value

                                sort_sheet_index = sort_sheet_index + 1

                        today = date.today()
                        dt_today = today.strftime("%d-%b-%Y")
                        wb_sorted_obj.save(sorted_seva_sheet)
                        print("Sorted sheet created for sorting")
                        self.obj_commonUtil.sortExcelSheetByDate(sorted_seva_sheet, sorted_seva_sheet)
                        now = datetime.now()
                        dt_string = now.strftime("%d_%b_%Y_%H%M%S")
                        currentyear = now.strftime("%Y")
                        destination_file = "..\\Expanse_Data\\" + currentyear + "\\Transaction\\Account_Statement\\Statements\\Gaushala_Account_Statement" + dt_string + ".pdf"
                        # write the  sorted record in statement template
                        template_sheet = "..\\Expanse_Data\\" + currentyear + "\\Transaction\\Account_Statement\\Template\\Transaction_template.xlsx"
                        wb_critical_stock = openpyxl.load_workbook(template_sheet)
                        critical_stock_sheet = wb_critical_stock.active
                        total_sorted_records = self.obj_commonUtil.totalrecords_excelDataBase(sorted_seva_sheet)
                        wb_sort = openpyxl.load_workbook(sorted_seva_sheet)
                        sort_sheet = wb_sort.active
                        dict_index = 1
                        starting_index = 15
                        print("Total sorted records :", total_sorted_records)
                        if total_sorted_records > 0:
                            text_info = "Statement is being generated for Main Accounts.Please wait ...."
                            infoLabel.configure(text=text_info, fg='purple')
                            for row_index in range(1, total_sorted_records + 1):
                                if dict_index == 1:
                                    # if credit column is numeric meaning transaction is credit candidate
                                    if str(sort_sheet.cell(row=row_index + 1, column=3).value).isnumeric():
                                        balance = int(
                                            sort_sheet.cell(row=row_index + 1, column=3).value)  # credit column
                                    else:
                                        # else transaction is debit candidate
                                        balance = int(
                                            sort_sheet.cell(row=row_index + 1, column=4).value)  # debit column
                                else:
                                    if str(sort_sheet.cell(row=row_index + 1,
                                                           column=3).value).isnumeric():  # credit amount is added
                                        balance = balance + int(sort_sheet.cell(row=row_index + 1, column=3).value)
                                    else:  # debit amount is subtracted
                                        balance = balance - int(sort_sheet.cell(row=row_index + 1, column=4).value)

                                for column_index in range(1, 7):
                                    if column_index == 1:  # Date
                                        text_value = sort_sheet.cell(row=row_index + 1, column=2).value
                                        text_value = text_value.strftime("%d-%b-%Y")
                                    elif column_index == 2:  # Invoice
                                        text_value = sort_sheet.cell(row=row_index + 1, column=10).value
                                    elif column_index == 3:  # Description
                                        text_value = str(sort_sheet.cell(row=row_index + 1, column=5).value) + "-By " + \
                                                     str(sort_sheet.cell(row=row_index + 1,
                                                                         column=6).value) + "-From " + \
                                                     str(sort_sheet.cell(row=row_index + 1, column=8).value)
                                    elif column_index == 4:  # credit
                                        text_value = sort_sheet.cell(row=row_index + 1, column=3).value
                                    elif column_index == 5:  # debit
                                        text_value = sort_sheet.cell(row=row_index + 1, column=4).value
                                    elif column_index == 6:  # balance
                                        text_value = str(balance)
                                    else:
                                        pass
                                    critical_stock_sheet.cell(row=starting_index, column=column_index).font = Font(
                                        size=8,
                                        name='Arial',
                                        bold=False)
                                    if column_index == 4 or column_index == 5 or column_index == 6:
                                        critical_stock_sheet.cell(row=starting_index,
                                                                  column=column_index).alignment = Alignment(
                                            horizontal='center', vertical='center', wrapText=True)
                                    else:
                                        critical_stock_sheet.cell(row=starting_index,
                                                                  column=column_index).alignment = Alignment(
                                            horizontal='left', vertical='center', wrapText=True)

                                    critical_stock_sheet.cell(row=starting_index,
                                                              column=column_index).value = text_value
                                dict_index = dict_index + 1
                                starting_index = starting_index + 1

                            frdateforstatement = fromDate.strftime("%d-%b-%Y")
                            todateforstatement = toDate.strftime("%d-%b-%Y")
                            critical_stock_sheet.cell(row=3, column=6).value = dt_today
                            critical_stock_sheet.cell(row=5, column=6).value = "Main Accounts"
                            critical_stock_sheet.cell(row=8, column=6).value = str(balance)
                            critical_stock_sheet.cell(row=9, column=6).value = str(starting_index - 15)
                            critical_stock_sheet.cell(row=10, column=6).value = frdateforstatement
                            critical_stock_sheet.cell(row=11, column=6).value = todateforstatement
                            wb_critical_stock.save(template_sheet)
                            print("File has been saved for template")
                            destination_copy_folder = InitDatabase.getInstance().get_desktop_statement_directory_path()
                            obj_threadClass = myThread(10, "statementThread", 1, template_sheet,
                                                       destination_file, starting_index, viewPDF, printBtn, infoLabel,
                                                       destination_copy_folder)
                            obj_threadClass.start()

                            print_result = partial(self.obj_commonUtil.open_statement_file,
                                                   template_sheet,
                                                   destination_file, starting_index)
                            printBtn.configure(command=print_result)
                            view_result = partial(self.obj_commonUtil.open_statement_file, template_sheet,
                                                  destination_file, starting_index)
                            viewPDF.configure(command=view_result, )

                            cancel_result = partial(self.closepage, template_sheet, starting_index,
                                                    gaushala_account_statement_window)
                            cancelbtn.configure(command=cancel_result)
                        else:
                            text_error = "No records present for Transactions in specified period!!!"
                            infoLabel.configure(text=text_error, fg='red')
                    else:
                        text_error = "No records present for Transactions in specified period!!!"
                        infoLabel.configure(text=text_error, fg='red')
        else:
            infoLabel.configure(text=error_info, fg='red')

    def initilize_sorted_Transaction_database(self):
        print("initilize_aorted_Transaction_database")
        today = datetime.now()
        year = today.strftime("%Y")
        dirname = "..\\Expanse_Data\\" + year + "\\Transaction\\Sorted_List"
        if not os.path.exists(dirname):
            print("Current year directory is not available , hence building one")
            os.makedirs(dirname)
        path = dirname + "\\Transaction.xlsx"
        self.sortedtransaction_database_name = path

        wb = openpyxl.Workbook()
        sheet = wb.active
        heading_list = ['S.No', 'Date', 'Credit Amt.(Rs.)', 'Debit Amt.(Rs)', 'Description', 'Mode of Transaction',
                        'Transaction By(Id)', 'Transaction By(Name)', 'Balance (Rs.)', 'Invoice ID']

        # preparing the top headers of the excel sheet.
        sheet.row_dimensions[1].height = 25

        excel_cell = {}
        alphabet_dict = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K']
        for iLoop in range(1, 11):
            sheet.cell(row=1, column=iLoop).font = Font(size=12, name='Times New Roman', bold=True)
            sheet.column_dimensions[alphabet_dict[iLoop - 1]].width = 22
            sheet.cell(row=1, column=iLoop).alignment = Alignment(horizontal='center', vertical='center')
            # sheet.cell(row=1, column=iLoop).fill = PatternFill(fgColor='20489F', fill_type='solid')
            excel_cell[iLoop] = sheet.cell(row=1, column=iLoop)
            excel_cell[iLoop].value = heading_list[iLoop - 1]
        sheet.column_dimensions['E'].width = 40
        wb.save(path)
        print(" Database for  Sorted Transaction initialized successfully ")

    def get_sorted_transaction_database_name(self):
        return self.sortedtransaction_database_name

    def initilize_new_directories(self):
        print("Initilize VYOAM Dir Structure")
        today = datetime.now()
        year = today.strftime("%Y")
        dirname1 = "Expanse_Data\\" + year + "\\Expanse"
        dirname2 = "Expanse_Data\\" + year + "\\Expanse\\Receipts"
        dirname3 = "Expanse_Data\\" + year + "\\Expanse\\Receipts\\Receipts"
        dirname4 = "Expanse_Data\\" + year + "\\Expanse\\Receipts\\Template"

        dirname5 = "Expanse_Data\\" + year + "\\Invoices"

        

        dirname18 = "Expanse_Data\\" + year + "\\Statements"

        dirname19 = "Expanse_Data\\" + year + "\\Transaction"
        dirname20 = "Expanse_Data\\" + year + "\\Transaction\\Account_Statement"
        dirname21 = "Expanse_Data\\" + year + "\\Transaction\\Account_Statement\\Statements"
        dirname22 = "Expanse_Data\\" + year + "\\Transaction\\Account_Statement\\Template"
        dirname23 = "Expanse_Data\\" + year + "\\Transaction\\Sorted_List"
        dirname24 = "Expanse_Data\\" + year + "\\Transaction\\StockSell"
        dirname25 = "Expanse_Data\\" + year + "\\Transaction\\StockSell\\Sorted"
        dirname26 = "Expanse_Data\\" + year + "\\Transaction\\StockSell\\Statements"
        dirname27 = "Expanse_Data\\" + year + "\\Transaction\\StockSell\\Template"

        directory_list = [dirname1, dirname2, dirname3, dirname4, dirname5, dirname18, dirname19, dirname20, dirname21,
                          dirname22, dirname23, dirname24, dirname25, dirname26, dirname27]

        for index in range(0, len(directory_list)):
            if not os.path.exists(directory_list[index]):
                print(directory_list[index], " ...created")
                os.makedirs(directory_list[index])

        print("VYOAM Dir Structure initialized")
        print("VYOAM Copying  common file start")
        shutil.copy("../Common_Files/Ashram_Voucher_Receipt_Template.xlsx", dirname4)
       
       

        shutil.copy("../Common_Files/Split_Donation_Open.xlsx", dirname22)
        shutil.copy("../Common_Files/Transaction_template.xlsx", dirname22)
        # shutil.copy("Common_Files\\stock_sales_account_template.xlsx", dirname27)
       

        # copy booklets , if they do not exists
       

        # copy booklets , if they do not exists
        filename_Ashram_Expanse_Voucher__booklet = dirname4 + "\\Ashram_Expanse_Voucher_Receipt_Booklet.xlsx"
        if not os.path.exists(filename_Ashram_Expanse_Voucher__booklet):
            print("Ashram_Expanse_Voucher__booklet new copied")
            shutil.copy("../Common_Files/Ashram_Expanse_Voucher_Receipt_Booklet.xlsx", dirname4)
        else:
            print("Ashram_Expanse_Voucher booklet already exist")



# Usage example:
if __name__ == "__main__":
    root = tk.Tk()
    app = LedgerForm(root)
    records = [
        ["2024-05-01", "Party A", "Location A", "INV001", "100.00", "80.00", "20.00", "MR001", "Remarks 1"],
        ["2024-05-02", "Party B", "Location B", "INV002", "200.00", "150.00", "50.00", "MR002", "Remarks 2"]
    ]
    app.show_records(records)
    root.mainloop()
    
